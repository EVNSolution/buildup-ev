# -*- coding: utf-8 -*-
"""하중계산서 생성기 — 원본 [최종] 4813 하중계산.pdf 픽셀 격자 재현.
사용: python3 gen_load_calc.py data.json out.xlsx
탈거/설치 항목은 개수에 따라 행이 가변(옵션 선택 결과에 연동).
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d = json.load(open(sys.argv[1], encoding="utf-8")) if len(sys.argv) > 1 else {}
OUT = sys.argv[2] if len(sys.argv) > 2 else "out.xlsx"
g = lambda *k: (d.get(k[0], {}) if len(k) == 1 else d.get(k[0], {}).get(k[1], "")) or ""
S = lambda v: "" if v in (None, "") else str(v)

INS = d.get("install_items", [])   # [{name, weight, dist}]
REM = d.get("remove_items", [])
CREW = d.get("crew", {})           # {name, persons, weight, dist}
NROW = max(1, len(INS), len(REM))  # ★ 가변 행 수

# ── 픽셀 격자 (원본 300dpi) ─────────────────────────────────
X=[83,246,379,454,507,565,728,779,1007,1026,1069,1268,1338,1378,1477,1603,1657,1775,1846,1934,2096,2115,2326]
Y=[125,257,340,411,482,553,623,694,765,836,907,978,1049,1135,1218,1288,1359,1430,1501,1572,1657,1740,1810,1881,1952,2023,2094,2168]
DROW = 59                      # 하단(탈거/설치/정원) 1행 높이 px
Y = Y + [2168 + DROW*(i+1) for i in range(NROW)]     # 헤더(2094-2168) 뒤로 가변 행 추가
NC, NR = len(X)-1, len(Y)-1

wb=Workbook(); ws=wb.active; ws.title="하중계산"
FN="맑은 고딕"; t=Side(style="thin",color="000000"); BD=Border(t,t,t,t)
F=lambda s,b=False: Font(name=FN,size=s,bold=b,color="000000")
AC=lambda: Alignment(horizontal="center",vertical="center",shrink_to_fit=True)
def P(r,c1,c2,txt="",sz=11,b=False,r2=None,box=True):
    r2=r2 or r
    if (r,c1)!=(r2,c2): ws.merge_cells(start_row=r,start_column=c1,end_row=r2,end_column=c2)
    if box:
        for rr in range(r,r2+1):
            for cc in range(c1,c2+1): ws.cell(rr,cc).border=BD
    x=ws.cell(r,c1); x.value=txt if txt!="" else None; x.font=F(sz,b); x.alignment=AC()

# ── 제목 / 제원 ────────────────────────────────────────────
P(1,1,NC,"하 중 계 산",sz=26,b=True)
P(2,1,4,""); P(2,5,7,"변경전"); P(2,8,10,"변경후")
P(2,11,15,""); P(2,16,19,"변경전"); P(2,20,22,"변경후")
sp=d.get("spec",{})
L=[("유형","type"),("승차정원(명)","seats"),("차량중량(kg)","curb"),("최대적재량(kg)","payload"),
   ("차량총중량(kg)","gvw"),("축간거리(mm)","wheelbase"),("하대옵셋트(mm)","offset")]
for i,(lb,k) in enumerate(L):
    r=3+i; P(r,1,4,lb); P(r,5,7,S(g("before",k)),sz=8 if i==0 else 11); P(r,8,10,S(g("after",k)),sz=8 if i==0 else 11)
for i,(lb,k) in enumerate([("전륜2륜간거리(mm)","tread2f"),("후륜2륜간거리(mm)","tread2r"),("조향륜하중분포율(%)","steer_ratio")]):
    r=10+i; P(r,1,4,lb,sz=10); P(r,5,10,S(g("after",k)))
P(3,11,12,"전체(mm)",r2=5)
for i,k in enumerate(["length","width","height"]):
    r=3+i; P(r,13,15,["길이","너비","높이"][i]); P(r,16,19,S(g("before",k))); P(r,20,22,S(g("after",k)))
P(6,11,12,"하대(mm)",r2=8)
for i,k in enumerate(["bed_len","bed_wid","bed_hgt"]):
    r=6+i; P(r,13,15,["길이","너비","높이"][i]); P(r,16,19,S(g("before",k))); P(r,20,22,S(g("after",k)))
P(9,11,15,""); P(9,16,19,"중량(kg)",sz=10); P(9,20,22,"후축까지의거리(mm)",sz=9)
ex=d.get("extra",{})
P(10,11,15,"최대적재량감소중량",sz=10); P(10,16,22,S(ex.get("payload_cut")))
P(11,11,15,"푸셔액슬/태그액슬",sz=10); P(11,16,19,S(ex.get("pusher_w"))); P(11,20,22,S(ex.get("pusher_d")))
P(12,11,15,"프레임절단중량",sz=10); P(12,16,19,S(ex.get("frame_w"))); P(12,20,22,S(ex.get("frame_d")))

# ── 축하중 분포 ────────────────────────────────────────────
P(13,1,3,"구분",r2=14); P(13,4,13,"공차시 하중분포"); P(13,14,22,"적차시 하중분포")
for c1,c2,tx in [(4,6,"튜닝전(kg)"),(7,8,"튜닝후(kg)"),(9,13,"타이어부하율(%)"),
                 (14,16,"튜닝전(kg)"),(17,19,"튜닝후(kg)"),(20,22,"타이어부하율(%)")]: P(14,c1,c2,tx,sz=11)
P(15,1,1,"전축",r2=16); P(17,1,1,"후축",r2=19)
AX=["전전","전후","후전","후중","후후"]
ld=d.get("load",{})
for i,ax in enumerate(AX):
    r=15+i; P(r,2,3,ax); v=ld.get(ax,{})
    for (c1,c2),k in zip([(4,6),(7,8),(9,13),(14,16),(17,19),(20,22)],
                         ["empty_before","empty_after","empty_ratio","full_before","full_after","full_ratio"]):
        P(r,c1,c2,S(v.get(k)))

# ── 타이어 ────────────────────────────────────────────────
P(20,1,3,"구분",r2=21); P(20,4,11,"튜닝전"); P(20,12,20,"튜닝후"); P(20,21,22,"바퀴수",r2=21)
P(21,4,9,"타이어형식"); P(21,10,11,"허용하중(kg)",sz=10); P(21,12,18,"타이어형식"); P(21,19,20,"허용하중(kg)",sz=10)
P(22,1,1,"전축",r2=23); P(24,1,1,"후축",r2=26)
tr=d.get("tire",{})
for i,ax in enumerate(AX):
    r=22+i; P(r,2,3,ax); v=tr.get(ax,{})
    for (c1,c2),k in zip([(4,9),(10,11),(12,18),(19,20),(21,22)],
                         ["spec_before","load_before","spec_after","load_after","wheels"]):
        P(r,c1,c2,S(v.get(k)))

# ── 하단: 설치/탈거/정원 (★가변 행) ──────────────────────────
H=27
P(H,1,2,"설치항목"); P(H,3,5,"중량"); P(H,6,7,"후축까지거리",sz=9)
P(H,8,10,"탈거항목"); P(H,11,11,"중량"); P(H,12,14,"후축까지거리",sz=9)
P(H,15,17,"항목"); P(H,18,19,"승차정원",sz=10); P(H,20,21,"중량"); P(H,22,22,"후축까지거리",sz=9)
for i in range(NROW):
    r=H+1+i
    if i < len(INS):
        it=INS[i]; P(r,1,2,S(it.get("name")),sz=6); P(r,3,5,S(it.get("weight"))); P(r,6,7,S(it.get("dist")))
    if i < len(REM):
        it=REM[i]; P(r,8,10,S(it.get("name")),sz=8); P(r,11,11,S(it.get("weight"))); P(r,12,14,S(it.get("dist")))
    if i == 0 and CREW:
        P(r,15,17,S(CREW.get("name"))); P(r,18,19,S(CREW.get("persons")))
        P(r,20,21,S(CREW.get("weight"))); P(r,22,22,S(CREW.get("dist")))

# ── 치수 (실측 보정식) ──────────────────────────────────────
for i in range(NC):
    ws.column_dimensions[get_column_letter(i+1)].width=round(max(0.3,(X[i+1]-X[i]+1.6)/23.51),2)
for i in range(NR):
    ws.row_dimensions[i+1].height=round((Y[i+1]-Y[i])*0.24*1.015,1)
ws.sheet_view.showGridLines=False
ws.page_setup.paperSize=9; ws.page_setup.orientation="portrait"; ws.page_setup.scale=100
ws.page_margins.left=83*25.4/300/25.4; ws.page_margins.right=0.35
ws.page_margins.top=125*25.4/300/25.4; ws.page_margins.bottom=0.3
ws.page_margins.header=0; ws.page_margins.footer=0
ws.print_area=f"A1:{get_column_letter(NC)}{NR}"
wb.save(OUT); print("saved", OUT, "| 탈거", len(REM), "설치", len(INS), "→", NROW, "행")
