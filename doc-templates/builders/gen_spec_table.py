# -*- coding: utf-8 -*-
"""주요제원대비표(별지 제33호의2서식) 생성기 — 원본 [최종] 주요제원 대비표.pdf 픽셀 격자 재현.
사용: python3 gen_spec_table.py data.json out.xlsx
격자·보정식은 40. 양식/_작업중_픽셀대조_진행상황.md 에서 확정된 값(수정 금지, 값 바인딩만).
값은 전부 data.json 에서 — 하드코딩 금지. 좌/우 블록 행 경계가 달라 기준행 = 좌·우 가로선 합집합.
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d = json.load(open(sys.argv[1], encoding="utf-8")) if len(sys.argv) > 1 else {}
OUT = sys.argv[2] if len(sys.argv) > 2 else "out.xlsx"
S = lambda v: "" if v in (None, "") else str(v)
def bi(sec, key, i):   # before/after 배열 값 (i=0 전, 1 후)
    v = d.get(sec, {}).get(key)
    return "" if v is None else (S(v[i]) if isinstance(v, list) else S(v))

# ── 격자 (300dpi px, 픽셀대조 확정값) ──────────────────────────────────────────
X = [242,328,378,431,474,547,620,766,1015,1265,1360,1406,1421,1468,1481,1511,1661,1812,1942,2025,2073,2240]
L = [358,459,531,604,676,748,821,882,942,1029,1106,1197,1287,1365,1442,1527,1604,1681,1776,1853,1930,2010,2076,2141,2206,2274,2351,2416,2481,2549,2636,2704,2790,2863,2938,3010,3087,3159,3235,3311]
R = [358,459,531,604,676,748,821,882,942,1029,1106,1182,1270,1365,1428,1492,1568,1644,1719,1783,1859,1935,2010,2076,2141,2206,2274,2351,2416,2481,2549,2636,2704,2790,2863,2938,3010,3159,3235,3311]
Y = sorted(set(L) | set(R))
IDX = {y: i for i, y in enumerate(Y)}
NC, NR = len(X) - 1, len(Y) - 1

def rb(y1, y2):   # y 픽셀구간 → (엑셀 시작행, 끝행) 1-base
    return IDX[y1] + 1, IDX[y2]

wb = Workbook(); ws = wb.active; ws.title = "주요제원대비표"
FN = "맑은 고딕"; t = Side(style="thin", color="000000"); BD = Border(t, t, t, t)
F = lambda s, b=False: Font(name=FN, size=s, bold=b, color="000000")
def AL(rot=0, wrap=False):
    return Alignment(horizontal="center", vertical="center", shrink_to_fit=not wrap,
                     wrap_text=wrap, text_rotation=rot)

def P(r1, c1, r2, c2, txt="", sz=8, b=False, rot=0, ha="center", wrap=False, box=True):
    if (r1, c1) != (r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    if box:
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(rr, cc).border = BD
    x = ws.cell(r1, c1); x.value = txt if txt != "" else None
    x.font = F(sz, b)
    a = AL(rot, wrap);
    if ha != "center": a = Alignment(horizontal=ha, vertical="center", shrink_to_fit=not wrap, wrap_text=wrap, text_rotation=rot)
    x.alignment = a

LB = True  # 라벨=bold 로 강조(원본 라벨 톤)

# ── 상단: 법령/제원관리번호/제목 (엑셀 밖 텍스트는 생략, 표 안에서 시작) ──────────
# 제목 (전폭)
r = rb(358, 459)
P(r[0], 1, r[1], NC, "튜닝 전ㆍ후 주요제원 대비표", sz=18, b=True)

# ── 1 소유자 주소 (값 전폭) ────────────────────────────────────────────────────
r = rb(459, 531)
P(r[0], 1, r[1], 1, "1", sz=8)
P(r[0], 2, r[1], 7, "소유자 주소", sz=8, b=LB)
P(r[0], 8, r[1], NC, S(d.get("owner_addr")), sz=8, ha="left")

# ── 2·3 / 4·5 / 6·7 / 8·9 (좌우 단순 라벨행) ──────────────────────────────────
def pair(y1, y2, ln, llbl, lval, rn, rlbl, rval):
    r = rb(y1, y2)
    P(r[0], 1, r[1], 1, ln); P(r[0], 2, r[1], 7, llbl, b=LB); P(r[0], 8, r[1], 9, lval)
    P(r[0], 10, r[1], 10, rn); P(r[0], 11, r[1], 17, rlbl, b=LB); P(r[0], 18, r[1], NC, rval)

pair(531, 604, "2", "성명",   S(d.get("owner_name")), "3", "최초등록일", S(d.get("reg_date")))
pair(604, 676, "4", "등록번호", S(d.get("reg_no")),    "5", "종별",     S(d.get("category")))
pair(676, 748, "6", "차명",   S(d.get("car_name")),   "7", "구분",     S(d.get("division")))
pair(748, 821, "8", "형식",   S(d.get("form_code")),  "9", "차체형상",  S(d.get("body_shape")))

# ── 10 승차정원 / 11 유형 (헤더행 + 값행) ─────────────────────────────────────
def hv(y_lbl_top, y_split, y_bot, num, lbl, sec, key, side):
    """side='L' → 번호 c1, 라벨 c2-7, 값 c8/c9. side='R' → c10, c11-17, c18-19/c20-21"""
    if side == 'L': nc, lc1, lc2, v1a, v1b, v2a, v2b = 1, 2, 7, 8, 8, 9, 9
    else:           nc, lc1, lc2, v1a, v1b, v2a, v2b = 10, 11, 17, 18, 19, 20, 21
    rt = rb(y_lbl_top, y_bot)   # 라벨은 헤더+값 2행 세로병합
    P(rt[0], nc, rt[1], nc, num); P(rt[0], lc1, rt[1], lc2, lbl, b=LB)
    rh = rb(y_lbl_top, y_split); rvv = rb(y_split, y_bot)
    P(rh[0], v1a, rh[1], v1b, "튜닝 전", b=LB); P(rh[0], v2a, rh[1], v2b, "튜닝 후", b=LB)
    P(rvv[0], v1a, rvv[1], v1b, bi(sec, key, 0)); P(rvv[0], v2a, rvv[1], v2b, bi(sec, key, 1))

hv(821, 882, 942, "10", "승차정원(명)", "val", "seats", 'L')
hv(821, 882, 942, "11", "유형",       "val", "type",  'R')

# ── 좌: 12 차량중량 / 14 최대적재량 / 15 차량총중량 ───────────────────────────
def lrow(y1, y2, num, lbl, sec, key, sz=8):
    r = rb(y1, y2)
    P(r[0], 1, r[1], 1, num); P(r[0], 2, r[1], 7, lbl, sz=sz, b=LB)
    P(r[0], 8, r[1], 8, bi(sec, key, 0)); P(r[0], 9, r[1], 9, bi(sec, key, 1))

lrow(942, 1029, "12", "차량중량(㎏)", "val", "curb")
lrow(1029, 1106, "14", "최대적재량(㎏)", "val", "payload")
lrow(1106, 1197, "15", "차량총중량(㎏)", "val", "gvw")

# ── 좌: 13 용도 (우측 942-1029, 헤더 없이 값행) ─────────────────────────────
r = rb(942, 1029)
P(r[0], 10, r[1], 10, "13"); P(r[0], 11, r[1], 17, "용도", b=LB)
P(r[0], 18, r[1], 19, bi("val", "usage", 0)); P(r[0], 20, r[1], 21, bi("val", "usage", 1))

# ── 좌: 16 차체제원 / 17 하대내측지수 (대라벨 c2-5, 소 c6-7, 값 c8/c9) ─────────
def dim3(y_top, num, lbl, sec, bands):
    rt = rb(y_top, bands[-1][1])
    P(rt[0], 2, rt[1], 5, lbl, sz=8, b=LB, wrap=True)
    P(rt[0], 1, rt[1], 1, num)
    for sub, (ya, yb) in zip(["길이", "너비", "높이"], bands):
        r = rb(ya, yb)
        P(r[0], 6, r[1], 7, sub, b=LB)
        P(r[0], 8, r[1], 8, bi(sec, {"길이":"len","너비":"wid","높이":"hgt"}[sub], 0))
        P(r[0], 9, r[1], 9, bi(sec, {"길이":"len","너비":"wid","높이":"hgt"}[sub], 1))

dim3(1197, "16", "차체제원(㎜)", "body", [(1197,1287),(1287,1365),(1365,1442)])
dim3(1442, "17", "하대내측지수(㎜)", "bed", [(1442,1527),(1527,1604),(1604,1681)])

# ── 좌: 18 하대옵셋트 ─────────────────────────────────────────────────────────
lrow(1681, 1776, "18", "하대옵셋트(㎜)", "val", "offset")

# ── 좌: 19 공차 / 20 적차 / 21 타이어하중율 (대 c2-4, 중 전/후 c5-6, 축 c7, 값 c8/c9)
# 5행: 전-전축, 전-후축, 후-전축, 후-중축, 후-후축
AX5 = [("전","전축","ff"), ("전","후축","fr"), ("후","전축","rf"), ("후","중축","rm"), ("후","후축","rr")]
def dist5(y_top, num, lbl, sec, bands):
    rt = rb(y_top, bands[-1][1])
    P(rt[0], 1, rt[1], 1, num); P(rt[0], 2, rt[1], 4, lbl, sz=8, b=LB, wrap=True)
    # 중(전/후) 세로병합: 전=처음2행, 후=나머지3행
    rf = rb(bands[0][0], bands[1][1]); rr = rb(bands[2][0], bands[4][1])
    P(rf[0], 5, rf[1], 6, "전", b=LB); P(rr[0], 5, rr[1], 6, "후", b=LB)
    for (grp, ax, k), (ya, yb) in zip(AX5, bands):
        r = rb(ya, yb)
        P(r[0], 7, r[1], 7, ax, b=LB)
        P(r[0], 8, r[1], 8, bi(sec, k, 0)); P(r[0], 9, r[1], 9, bi(sec, k, 1))

dist5(1776, "19", "공차 시 하중분포(㎏)", "empty",
      [(1776,1853),(1853,1930),(1930,2010),(2010,2076),(2076,2141)])
dist5(2141, "20", "적차 시 하중분포(㎏)", "loaded",
      [(2141,2206),(2206,2274),(2274,2351),(2351,2416),(2416,2481)])
dist5(2481, "21", "적재 시 타이어 하중율(%)", "tire_rate",
      [(2481,2549),(2549,2636),(2636,2704),(2704,2790),(2790,2863)])

# ── 좌: 22 앞바퀴분포율 / 23 축간거리 / 24 제작허용총중량 / 25 모델연도 ─────────
r = rb(2863, 2938)
P(r[0],1,r[1],1,"22"); P(r[0],2,r[1],7,"적재 시 앞바퀴 하중분포율(%)", sz=7, b=LB, wrap=True)
P(r[0],8,r[1],8, bi("val","front_ratio",0)); P(r[0],9,r[1],9, bi("val","front_ratio",1))
r = rb(2938, 3010)
P(r[0],1,r[1],1,"23"); P(r[0],2,r[1],7,"축간거리(㎜)", b=LB)
P(r[0],8,r[1],8, S(d.get("wheelbase"))); P(r[0],9,r[1],9, S(d.get("wheelbase")))
r = rb(3010, 3087)
P(r[0],1,r[1],1,"24"); P(r[0],2,r[1],7,"제작허용총중량(㎏)", sz=8, b=LB)
P(r[0],8,r[1],9, S(d.get("gvwr")))   # 값 전/후 병합
r = rb(3087, 3159)
P(r[0],1,r[1],1,"25"); P(r[0],2,r[1],7,"모델연도", b=LB)
P(r[0],8,r[1],9, S(d.get("model_year")))

# ── 우: 26 원동기 (1029-2076, col10 세로 "원동기", 소구분 c11-17, 값 c18-19/c20-21)
r26 = rb(1029, 2076)
P(r26[0], 10, r26[1], 10, "26\n원동기", sz=8, b=LB, rot=90)
eng = "engine"
def erow(ya, yb, lbl, key, big=None, big_span=None, big_col=(11,12), lblcol=(13,17)):
    r = rb(ya, yb)
    if big and big_span:
        rbig = rb(*big_span)
        P(rbig[0], big_col[0], rbig[1], big_col[1], big, sz=7, b=LB, rot=90, wrap=True)
    P(r[0], lblcol[0], r[1], lblcol[1], lbl, sz=7, b=LB, wrap=True)
    P(r[0], 18, r[1], 19, bi(eng, key, 0)); P(r[0], 20, r[1], 21, bi(eng, key, 1))

# 사용연료 (라벨 c13-17 → 실제 사용연료는 중+소 병합; 대라벨칸 c11-12 비움)
r = rb(1029, 1106); P(r[0],11,r[1],12,"", b=False); P(r[0],13,r[1],17,"사용연료", sz=7, b=LB)
P(r[0],18,r[1],19, bi("val","fuel",0)); P(r[0],20,r[1],21, bi("val","fuel",1))
# 내연기관 (대라벨 세로 c11-12, 3행)
erow(1106,1182,"형식","ice_form", big="내연기관", big_span=(1106,1365))
erow(1182,1270,"최고출력(ps/rpm)","ice_power")
erow(1270,1365,"기통수/총배기량(cc)","ice_cyl")
# 구동전동기 (3행)
erow(1365,1428,"형식","mot_form", big="구동전동기", big_span=(1365,1568))
erow(1428,1492,"개수","mot_cnt")
erow(1492,1568,"최고출력(KW/rpm)","mot_power")
# 구동축전지 (3행)
erow(1568,1644,"정격전압/용량(V/Ah)","bat_volt", big="구동축전지", big_span=(1568,1783))
erow(1644,1719,"형식","bat_form")
erow(1719,1783,"개수","bat_cnt")
# 연료전지 (3행)
erow(1783,1859,"형식","fc_form", big="연료전지", big_span=(1783,2010))
erow(1859,1935,"개수","fc_cnt")
erow(1935,2010,"최고출력(KW)","fc_power")
# 하이브리드 시스템 (중+소 병합)
r = rb(2010,2076); P(r[0],11,r[1],12,"", b=True); P(r[0],13,r[1],17,"하이브리드 시스템", sz=7, b=LB)
P(r[0],18,r[1],19, bi(eng,"hybrid",0)); P(r[0],20,r[1],21, bi(eng,"hybrid",1))

# ── 우: 27 연료소비율 (2076-2274, 항목\모드 사선 + 복합 c18 / 시가지 c19-20 / 고속 c21)
r27 = rb(2076, 2274)
P(r27[0], 10, r27[1], 10, "27", sz=8)
P(r27[0], 11, r27[1], 12, "연료\n소비율", sz=7, b=LB, rot=90, wrap=True)
r = rb(2076, 2141)   # 헤더행: 항목\모드
P(r[0],13,r[1],17,"항목＼모드", sz=7, b=LB)
P(r[0],18,r[1],18,"복합", sz=7, b=LB); P(r[0],19,r[1],20,"시가지", sz=7, b=LB); P(r[0],21,r[1],21,"고속", sz=7, b=LB)
econ = d.get("econ", {})
def econrow(ya, yb, lbl, key):
    r = rb(ya, yb); v = econ.get(key, ["","",""])
    P(r[0],13,r[1],17,lbl, sz=6, b=LB, wrap=True)
    P(r[0],18,r[1],18,S(v[0])); P(r[0],19,r[1],20,S(v[1])); P(r[0],21,r[1],21,S(v[2]))
econrow(2141,2206,"연비(㎞/ℓ,㎞/kwh,km/kg)","eff")
econrow(2206,2274,"1회충전주행거리(㎞)","range")

# ── 우: 28 윤간거리 / 29 변속기 / 31 구동방식 / 32 차대번호 ──────────────────
def rrow2(y1, ymid, y2, num, lbl, k1lbl, k1, k2lbl, k2, sec="misc"):
    """2행 그룹: 라벨 세로병합, 소라벨 c13-17, 값 c18-19/c20-21"""
    rt = rb(y1, y2)
    P(rt[0],10,rt[1],10,num); P(rt[0],11,rt[1],12,lbl, sz=7, b=LB, rot=90, wrap=True)
    for (ya,yb,sl,kk) in [(y1,ymid,k1lbl,k1),(ymid,y2,k2lbl,k2)]:
        r = rb(ya,yb); P(r[0],13,r[1],17,sl, sz=7, b=LB)
        P(r[0],18,r[1],19, bi(sec,kk,0)); P(r[0],20,r[1],21, bi(sec,kk,1))

rrow2(2274,2351,2416,"28","윤간거리(㎜)","전","tread_f","후","tread_r","misc")
rrow2(2416,2481,2549,"29","변속기","종류","trans_type","단수","trans_speed","misc")

# ── 우: 30 타이어 형식 (2549-2938, 대 c11-15, 전/후 c16, 축 c17, 값 c18-19/c20-21)
r30 = rb(2549, 2938)
P(r30[0],10,r30[1],10,"30"); P(r30[0],11,r30[1],15,"타이어 형식", sz=7, b=LB, rot=90, wrap=True)
TIRE5 = [("전","전축","f_front"),("전","후축","f_rear"),("후","전축","r_front"),("후","중축","r_mid"),("후","후축","r_rear")]
TBANDS = [(2549,2636),(2636,2704),(2704,2790),(2790,2863),(2863,2938)]
rf = rb(2549,2704); rr = rb(2704,2938)
P(rf[0],16,rf[1],16,"전", b=LB); P(rr[0],16,rr[1],16,"후", b=LB)
for (grp,ax,k),(ya,yb) in zip(TIRE5, TBANDS):
    r = rb(ya,yb); P(r[0],17,r[1],17,ax, sz=7, b=LB)
    P(r[0],18,r[1],19, bi("tire",k,0)); P(r[0],20,r[1],21, bi("tire",k,1))

# 31 구동방식
r = rb(2938,3010); P(r[0],10,r[1],10,"31"); P(r[0],11,r[1],17,"구동방식", b=LB)
P(r[0],18,r[1],19, bi("val","drivetrain",0)); P(r[0],20,r[1],21, bi("val","drivetrain",1))
# 32 차대번호 (좌 24·25 두 행 높이 = 3010-3159, 값 전폭 병합)
r = rb(3010,3159); P(r[0],10,r[1],10,"32"); P(r[0],11,r[1],17,"차대번호(1~8자리)", sz=7, b=LB)
P(r[0],18,r[1],21, S(d.get("vin")))

# ── 33 튜닝개요 / 34 비고 (전폭) ─────────────────────────────────────────────
r = rb(3159,3235); P(r[0],1,r[1],1,"33"); P(r[0],2,r[1],7,"튜닝 개요", b=LB)
P(r[0],8,r[1],NC, S(d.get("tuning_summary")), ha="left")
r = rb(3235,3311); P(r[0],1,r[1],1,"34"); P(r[0],2,r[1],7,"비고", b=LB)
P(r[0],8,r[1],NC, S(d.get("remark")), ha="left")

# ── 치수 (실측 보정식, 하중계산서와 동일) ────────────────────────────────────
for i in range(NC):
    ws.column_dimensions[get_column_letter(i+1)].width = round(max(0.3, (X[i+1]-X[i]+1.6)/23.51), 2)
for i in range(NR):
    ws.row_dimensions[i+1].height = round((Y[i+1]-Y[i])*0.24*1.015, 1)
ws.sheet_view.showGridLines = False
ws.page_setup.paperSize = 9; ws.page_setup.orientation = "portrait"; ws.page_setup.scale = 100
ws.page_margins.left = X[0]/300; ws.page_margins.right = 0.2
ws.page_margins.top = Y[0]/300; ws.page_margins.bottom = 0.2
ws.page_margins.header = 0; ws.page_margins.footer = 0
ws.print_area = f"A1:{get_column_letter(NC)}{NR}"
wb.save(OUT); print("saved", OUT, "| rows", NR, "cols", NC)
