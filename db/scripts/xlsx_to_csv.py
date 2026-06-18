#!/usr/bin/env python3
"""
db/templates/*.xlsx → db/seed/{table_name}.csv 변환 스크립트

실행: python3 db/scripts/xlsx_to_csv.py

규칙:
- 행 1: 시트 제목 (스킵)
- 행 2: 설명 (스킵)
- 행 3: 한글 컬럼명 (스킵)
- 행 4: 영문 컬럼명 → CSV 헤더 ([PK], [FK→X], [선택▼] 태그 제거)
- 행 5+: 데이터 (마지막 셀이 '← 예시'인 행 제외, 완전 빈 행 제외)
- 예시만 있는 시트: 헤더만 담은 빈 CSV 생성
"""
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 필요: pip3 install openpyxl")

REPO_ROOT = Path(__file__).parent.parent.parent
TEMPLATE = REPO_ROOT / "db" / "templates" / "buildup-ev_DB구축템플릿_v0.2.xlsx"
SEED_DIR = REPO_ROOT / "db" / "seed"

SKIP_SHEETS = {"0_가이드라인", "E_거래(시스템자동)"}


def extract_key(cell_val: object) -> str | None:
    """'[PK] code · VARCHAR · 필수' → 'code'"""
    if not cell_val:
        return None
    s = re.sub(r"\[.*?\]\s*", "", str(cell_val))
    s = s.split("·")[0].strip()
    s = re.sub(r"[^a-zA-Z0-9_]", "", s)
    return s or None


def is_example_row(row: tuple) -> bool:
    vals = [v for v in row if v is not None]
    return bool(vals) and str(vals[-1]).strip() == "← 예시"


def table_name_from_title(title: str, sheet_name: str) -> str:
    m = re.search(r"\((\w+)\)", str(title))
    return m.group(1) if m else sheet_name


def convert():
    SEED_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(TEMPLATE, data_only=True)

    written = []
    skipped = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            skipped.append(sheet_name)
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            skipped.append(sheet_name)
            continue

        table = table_name_from_title(rows[0][0] if rows[0] else "", sheet_name)
        headers = [extract_key(c) for c in rows[3]]
        # None 헤더 열 위치 기록 (데이터 행 자르기용)
        valid_cols = [i for i, h in enumerate(headers) if h]
        clean_headers = [headers[i] for i in valid_cols]

        data_rows = []
        for row in rows[4:]:
            if all(c is None for c in row):
                continue
            if is_example_row(row):
                continue
            data_rows.append([row[i] if i < len(row) else None for i in valid_cols])

        out_path = SEED_DIR / f"{table}.csv"
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(clean_headers)
            writer.writerows(data_rows)

        status = f"{len(data_rows)}행" if data_rows else "헤더만(미입력)"
        written.append(f"  {out_path.name:35s} ← {sheet_name} ({status})")

    print(f"✓ {TEMPLATE.name} → db/seed/")
    for line in written:
        print(line)
    if skipped:
        print(f"\n스킵: {', '.join(skipped)}")


if __name__ == "__main__":
    convert()
